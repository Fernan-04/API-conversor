"""Lector XLSX: implementa el puerto `DocumentReader` (Fase 1).

Usa `openpyxl` en modo `read_only` + `data_only` (valores calculados, no
fórmulas). Cada hoja es una sección: un `Heading` con el nombre de la hoja y una
`Table` con sus filas. Se recortan las filas/columnas vacías al borde y se aplica
el tope `config.xlsx_max_rows` para hojas enormes.
"""

from __future__ import annotations

import io

from doc2md.adapters.outbound._table_clean import clean_table
from doc2md.adapters.outbound._zip_guard import ensure_safe_zip
from doc2md.config import Config
from doc2md.domain.errors import CorruptFileError
from doc2md.domain.models import Document, Element, Heading, Paragraph, Table
from doc2md.text_utils import clean_text


def _cell_to_str(value, config: Config) -> str:
    if value is None:
        return ""
    return clean_text(str(value), config)


class XlsxReader:
    """Adaptador de lectura XLSX (puerto `DocumentReader`)."""

    def read(self, data: bytes, filename: str, config: Config) -> Document:
        ensure_safe_zip(data, config)   # guarda anti zip-bomb (XLSX es zip+XML)
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise CorruptFileError(detail=str(exc)) from exc

        sections: list[list[Element]] = []
        try:
            for ws in wb.worksheets:
                elements: list[Element] = [Heading(level=1, text=ws.title or "Hoja")]
                rows: list[list[str]] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= config.xlsx_max_rows:
                        break
                    rows.append([_cell_to_str(c, config) for c in row])
                rows = clean_table(rows, config)
                if rows:
                    ncol = len(rows[0])
                    if ncol > config.xlsx_max_cols:
                        rows = [r[: config.xlsx_max_cols] for r in rows]
                        elements.append(Paragraph(
                            text=f"(Hoja recortada a {config.xlsx_max_cols} de "
                                 f"{ncol} columnas: parece una hoja de maquetación "
                                 f"o un diagrama de Gantt, no una tabla de datos.)"
                        ))
                    elements.append(Table(rows=rows))
                sections.append(elements)
        finally:
            wb.close()

        return Document(sections=sections)
