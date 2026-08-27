"""Tests de los lectores DOCX / PPTX / XLSX (Fase 1).

Las fixtures se generan al vuelo (ver conftest.py) porque no hay originales
reales de estos formatos en el repo.
"""

from __future__ import annotations

import io

from doc2md import convert
from doc2md.config import Config

from conftest import GANTT, requires_gantt


def _xlsx_from_rows(rows, title="Hoja"):
    """Construye un .xlsx en memoria a partir de filas (para los tests)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = title
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# DOCX
# --------------------------------------------------------------------------- #

def test_docx_headings(docx_bytes):
    md = convert(docx_bytes, Config(), filename="doc.docx")
    assert "# Título Principal" in md
    assert "## Sección Uno" in md


def test_docx_paragraph_and_bold(docx_bytes):
    md = convert(docx_bytes, Config(), filename="doc.docx")
    assert "Este es un párrafo normal de introducción." in md
    assert "**Etiqueta importante**" in md


def test_docx_list(docx_bytes):
    md = convert(docx_bytes, Config(), filename="doc.docx")
    assert "- Primer ítem" in md
    assert "- Segundo ítem" in md


def test_docx_inline_bold_italic(docx_bytes):
    """Negrita y cursiva PARCIALES dentro de un párrafo (no todo el párrafo)."""
    md = convert(docx_bytes, Config(), filename="doc.docx")
    assert "Texto con **negrita** y *cursiva*." in md


def test_docx_hyperlink(docx_bytes):
    """Los hipervínculos se conservan como [texto](url) (antes se perdían)."""
    md = convert(docx_bytes, Config(), filename="doc.docx")
    assert "[nuestra web](https://example.com/x)" in md


def test_docx_numbered_list(docx_bytes):
    """Una lista con estilo 'List Number' se emite con marcador numerado."""
    md = convert(docx_bytes, Config(), filename="doc.docx")
    assert "1. Paso uno" in md
    assert "1. Paso dos" in md


def test_docx_table(docx_bytes):
    md = convert(docx_bytes, Config(), filename="doc.docx")
    assert "| A | B | C |" in md
    assert "| 1 | 2 | 3 |" in md


# --------------------------------------------------------------------------- #
# PPTX
# --------------------------------------------------------------------------- #

def test_pptx_title_as_heading(pptx_bytes):
    md = convert(pptx_bytes, Config(), filename="pres.pptx")
    assert "# Diapositiva de Prueba" in md


def test_pptx_bullets_and_notes(pptx_bytes):
    md = convert(pptx_bytes, Config(), filename="pres.pptx")
    assert "- Punto uno" in md
    assert "- Punto dos" in md
    # Notas del orador como bloque aparte.
    assert "Notas" in md
    assert "Nota del orador." in md


def test_pptx_table(pptx_bytes):
    md = convert(pptx_bytes, Config(), filename="pres.pptx")
    assert "| Col1 | Col2 |" in md
    assert "| x | y |" in md


def test_pptx_title_not_duplicated(pptx_bytes):
    """El título de la diapositiva sale UNA vez (como #), no repetido como párrafo.

    Regresión: python-pptx devuelve un wrapper distinto en cada acceso, así que
    comparar el shape del título con `is` fallaba y lo emitía dos veces.
    """
    md = convert(pptx_bytes, Config(), filename="pres.pptx")
    assert md.count("Diapositiva de Prueba") == 1
    assert "# Diapositiva de Prueba" in md


# --------------------------------------------------------------------------- #
# XLSX
# --------------------------------------------------------------------------- #

def test_xlsx_sheets_as_sections(xlsx_bytes):
    md = convert(xlsx_bytes, Config(), filename="libro.xlsx")
    assert "# Ventas" in md
    assert "# Resumen" in md


def test_xlsx_table_columns(xlsx_bytes):
    md = convert(xlsx_bytes, Config(), filename="libro.xlsx")
    assert "| Producto | Cantidad | Precio |" in md
    assert "| Manzana | 10 | 5 |" in md


def test_xlsx_row_limit():
    """El tope de filas recorta hojas enormes."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    for i in range(50):
        ws.append([f"fila{i}", i])
    buf = io.BytesIO()
    wb.save(buf)

    md = convert(buf.getvalue(), Config(xlsx_max_rows=10), filename="big.xlsx")
    assert "fila9" in md
    assert "fila20" not in md


def test_xlsx_wide_sheet_truncated():
    """Una hoja muy ancha (tipo Gantt) se recorta a xlsx_max_cols con aviso."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    ws.append([f"c{i}" for i in range(200)])   # 200 columnas de datos
    ws.append([str(i) for i in range(200)])
    buf = io.BytesIO()
    wb.save(buf)

    md = convert(buf.getvalue(), Config(xlsx_max_cols=20), filename="gantt.xlsx")
    assert "recortada a 20 de 200 columnas" in md
    # La cabecera de la tabla queda en 20 columnas (20 celdas + 2 bordes vacíos).
    header = next(l for l in md.splitlines() if l.startswith("| c0 "))
    assert header.count("|") == 21


def test_xlsx_empty_columns_dropped():
    """Las columnas totalmente vacías intercaladas se descartan."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    ws.append(["A", None, None, "B"])     # dos columnas vacías en medio
    ws.append(["1", None, None, "2"])
    buf = io.BytesIO()
    wb.save(buf)

    md = convert(buf.getvalue(), Config(), filename="huecos.xlsx")
    assert "| A | B |" in md
    assert "| 1 | 2 |" in md


# --------------------------------------------------------------------------- #
# XLSX — limpieza de hojas de maquetación (Gantt/cronograma): quitar el ruido
# --------------------------------------------------------------------------- #

def test_xlsx_error_values_normalized():
    """Los errores de fórmula (#N/A, #REF!…) se tratan como celda vacía."""
    data = _xlsx_from_rows([
        ["Nombre", "Valor", "Extra"],
        ["Ana", 10, "#N/A"],
        ["Beto", "#REF!", "ok"],
    ])
    md = convert(data, Config(), filename="errores.xlsx")
    assert "#N/A" not in md and "#REF!" not in md
    assert "Ana" in md and "Beto" in md and "ok" in md


def test_xlsx_dead_tail_trimmed():
    """La cola muerta (solo un índice incremental / ceros) se recorta."""
    rows = [["Actividad", "Responsable", "Estado"],
            ["Diseño", "Alumno 1", "Completado"],
            ["Pruebas", "Alumno 2", "En Proceso"]]
    # 40 filas de andamiaje: solo un contador en la 1ª columna, resto vacío.
    for i in range(1, 41):
        rows.append([i, None, None])
    md = convert(_xlsx_from_rows(rows), Config(), filename="cola.xlsx")
    assert "Diseño" in md and "En Proceso" in md      # la tabla real sobrevive
    assert md.count("\n| ") < 8                          # la cola muerta se fue


def test_xlsx_single_int_column_preserved():
    """Salvaguarda: una lista de una sola columna de enteros NO se borra como cola."""
    data = _xlsx_from_rows([[i] for i in range(1, 31)])
    md = convert(data, Config(), filename="ids.xlsx")
    assert "| 30 |" in md and "| 1 |" in md              # se conservan todos


def test_xlsx_layout_timeline_columns_dropped():
    """Una hoja ancha tipo Gantt pierde la banda de columnas de línea de tiempo
    (casi vacías), pero conserva la tabla real de la izquierda."""
    # Cabecera SIN etiquetas de banda (como una línea de tiempo sin encabezado
    # denso) y 20 filas de datos donde la banda de 40 columnas está casi vacía:
    # cada columna recibe a lo sumo una marca -> fill << 0.08 -> se descarta.
    rows = [["Actividad", "Responsable", "Estado"] + [""] * 40]
    for a in range(40):
        row = [f"Act {a}", f"Alumno {a % 3}", "Completado"] + [""] * 40
        row[3 + a] = "x"            # una marca por columna de la banda (fill ~1/41)
        rows.append(row)
    md = convert(_xlsx_from_rows(rows), Config(), filename="gantt.xlsx")
    assert "Act 0" in md and "Completado" in md
    # La cabecera de la tabla no debe arrastrar las 40 columnas de la banda.
    header_line = next(l for l in md.splitlines() if l.startswith("| Actividad "))
    assert header_line.count("|") <= 8                   # ~3-4 columnas reales


def test_xlsx_normal_sheet_unchanged():
    """Regresión: una hoja de datos normal (angosta y llena) queda intacta."""
    data = _xlsx_from_rows([
        ["Producto", "Cantidad", "Precio"],
        ["Manzana", 10, 5],
        ["Pera", 3, 8],
    ], title="Datos")
    md = convert(data, Config(), filename="datos.xlsx")
    assert "| Producto | Cantidad | Precio |" in md
    assert "| Manzana | 10 | 5 |" in md
    assert "| Pera | 3 | 8 |" in md


@requires_gantt
def test_xlsx_real_gantt_cleaned():
    """El Gantt real se reduce drásticamente sin perder los datos valiosos."""
    md = convert(str(GANTT))
    assert len(md) < 15000                               # hoy sin limpiar: ~42 KB
    assert md.count("\n") < 200                           # hoy sin limpiar: ~580
    for label in ("Completado", "Alumno 1", "Alumno 3", "FASE 1", "ESTADO", "AVANCE"):
        assert label in md, f"se perdió un dato real: {label}"
